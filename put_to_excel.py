from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook, styles, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from count_params import light_avg_temp, change_pressure

FILENAME = "Weather.xlsx"

COL_TEMP = 1
COL_PRESSURE = 2
COL_HUMIDITY = 3
COL_CONDITION = 4
COL_DATE = 1

STRING_TITLE = 3
STRING_MORNING = 4
STRING_DAY = 5
STRING_EVENING = 6
STRING_NIGHT = 7
STRING_AVG_TEMP = 8
STRING_MAGNETIC = 9
STRING_CHANGE_PRESSURE = 10

SHARP_INCREASE = 'ожидается резкое увеличение атмосферного давления'
SHARP_DECREASE = 'ожидается резкое падение атмосферного давления'

PERIOD = {0: STRING_MORNING, 1: STRING_DAY, 2: STRING_EVENING, 3: STRING_NIGHT}


def create_new_file():
    wb = Workbook()
    ws = wb.active
    ws.insert_cols(1, 5)
    ws.parent.save(FILENAME)
    return ws


def create_title(ws, i):
    ws[STRING_TITLE + i][COL_TEMP].value = 'Температура, C'
    ws[STRING_TITLE + i][COL_PRESSURE].value = 'Давление'
    ws[STRING_TITLE + i][COL_HUMIDITY].value = 'Влажность'
    ws[STRING_TITLE + i][COL_CONDITION].value = 'Погодное явление'
    ws[STRING_MORNING + i][0].value = 'Утром'
    ws[STRING_DAY + i][0].value = 'Днём'
    ws[STRING_EVENING + i][0].value = 'Вечером'
    ws[STRING_NIGHT + i][0].value = 'Ночью'
    ws[STRING_AVG_TEMP + i][0].value = 'Ср. темп. светового дня'
    ws[STRING_CHANGE_PRESSURE + i][0].value = 'Резкое изменение давления'
    ws[STRING_MAGNETIC + i][0].value = 'Магнитное поле'


def write_to_excel(arr_: list, city: str) -> None:
    ws = create_new_file()
    i = 1
    ws.insert_rows(12)
    ws[1][0].value = f'Прогноз погоды на 7 дней для города: {city}'
    for day in arr_:
        ws[1 + i][0].value = f'Погода на {day["date"]}'
        create_title(ws, i)
        for key in PERIOD.keys():
            time_of_day = day["periods"][key]
            ws[PERIOD[key] + i][COL_TEMP].value = f'{time_of_day["temp"]["min"]} ... {time_of_day["temp"]["max"]}'
            ws[PERIOD[key] + i][COL_PRESSURE].value = time_of_day["pressure"]
            ws[PERIOD[key] + i][COL_HUMIDITY].value = time_of_day["humidity"]
            ws[PERIOD[key] + i][COL_CONDITION].value = time_of_day["condition"]
        ws[STRING_AVG_TEMP + i][1].value = light_avg_temp(day['periods'])
        ws[STRING_MAGNETIC + i][1].value = day['magnetic field']
        pressure = change_pressure(day['periods'])
        if pressure[0]:
            ws[STRING_CHANGE_PRESSURE + i][1].value = SHARP_INCREASE if pressure[1] == 1 else SHARP_DECREASE

        ws.insert_rows(12)
        i += 14

    ws.parent.save(FILENAME)
